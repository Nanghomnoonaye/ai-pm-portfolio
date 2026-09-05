from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb=Workbook()
F="Arial"; hdr=Font(name=F,bold=True,color="FFFFFF"); hf=PatternFill("solid",fgColor="4F81BD")
def sheet(name, headers, rows, widths=None):
    ws=wb.create_sheet(name)
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.font=hdr; cell.fill=hf; cell.alignment=Alignment(wrap_text=True,vertical="center")
    for r,row in enumerate(rows,2):
        for c,v in enumerate(row,1):
            cell=ws.cell(r,c,v); cell.font=Font(name=F)
    for c in range(1,len(headers)+1):
        ws.column_dimensions[chr(64+c)].width=(widths[c-1] if widths else 18)
    ws.freeze_panes="A2"
    return ws
wb.remove(wb.active)

# 1) Skills reference
sheet("Skills",
 ["Skill","Description","Category"],
 [["AI Literacy","Understand bias, model drift, data training","Technical"],
  ["Data Strategy","Data quality, availability, governance","Data"],
  ["Responsible AI","Fairness, transparency, privacy, ethics","Governance"],
  ["Cross-Functional Collaboration","Work with data scientists, engineers, stakeholders","People"],
  ["Experimentation Mindset","A/B testing, iteration, hypothesis-led development","Method"]],
 widths=[30,45,14])

# 2) AI-PM examples: tidy data with numeric scores (1-5) for analysis
sheet("AI_PM_Examples",
 ["Industry","Example","PM_Focus","Business_Impact","Data_Maturity",
  "AI_Literacy","Data_Strategy","Responsible_AI","Cross_Functional","Experimentation"],
 [["Retail","Personalized recommendation engine","Conversion / CX",5,5,4,5,3,4,5],
  ["Manufacturing","Predictive maintenance","Uptime / cost",4,4,5,4,2,5,4],
  ["Healthcare","CGM AI coaching for Type 2 diabetes","Outcomes / HbA1c",5,3,5,4,5,5,3],
  ["Finance","Fraud detection","Risk / loss prevention",5,5,4,5,4,4,4],
  ["Transportation","Route & demand optimization","Efficiency / ETA",4,4,4,3,3,4,4],
  ["Government","Public-health risk prediction","Population health",4,3,3,5,5,4,3]],
 widths=[16,34,20,15,14,12,12,14,14,15])

# 3) General PM examples (from the course content)
sheet("PM_Examples",
 ["Example","Domain","Key_PM_Activities","Skill_Highlighted"],
 [["Launching a new smartphone model","Consumer tech",
   "Identify needs (battery, camera); align vision; coordinate eng/marketing/sales; report to execs","Cross-functional leadership"],
  ["Launching an in-app purchase feature","Mobile app",
   "Analyze user data; forecast ROI & payback; align to strategy; manage compliance","Business acumen"],
  ["AI recommendation engine","Retail",
   "Build model with data scientists; A/B test algorithms; monitor bias; ensure transparency","AI literacy + data strategy"],
  ["Predictive maintenance","Manufacturing",
   "Analyze sensor data; predict failures; schedule maintenance; adapt roadmap to model cycles","Data strategy"],
  ["CGM AI coaching","Healthcare",
   "Personalize glucose coaching; validate clinically; guard privacy (PDPA); clinician override","Responsible AI"]],
 widths=[34,16,55,26])

# 4) Experimentation steps
sheet("Experimentation",
 ["Step","Description"],
 [["Design A/B tests","Compare model/feature versions in controlled experiments"],
  ["Hypothesis-driven development","State expected impact before testing"],
  ["Iterative improvement","Refine models from results (bias, accuracy)"],
  ["Monitor metrics","Track KPIs to evaluate impact and guide decisions"],
  ["Culture of learning","Promote data-driven, open experimentation"]],
 widths=[30,60])

wb.save("data/AI_PM_Examples.xlsx"); print("xlsx saved")
